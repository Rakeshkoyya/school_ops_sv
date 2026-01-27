"""Authentication service."""

from io import BytesIO
from datetime import datetime, timezone
from uuid import UUID

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.core.exceptions import AuthenticationError, NotFoundError, ValidationError
from app.core.security import (
    create_access_token,
    create_refresh_token,
    hash_password,
    verify_password,
    verify_refresh_token,
)
from app.models.user import User
from app.models.rbac import Role, UserRoleProject
from app.models.project import Project
from app.schemas.auth import (
    AdminUserUpdate,
    LoginRequest,
    ProjectRoleMapping,
    ProjectUserUpdate,
    RoleInfo,
    TokenResponse,
    UserBulkUploadResult,
    UserCreate,
    UserResponse,
    UserWithProjectRoles,
)


# Excel template columns for user upload
USER_TEMPLATE_COLUMNS = [
    ("name", "Full Name", True),
    ("username", "Username", True),
    ("phone", "Phone", False),
    ("password", "Password", True),
]


class AuthService:
    """Authentication service."""

    def __init__(self, db: Session):
        self.db = db

    def login(self, request: LoginRequest) -> TokenResponse:
        """Authenticate user and return tokens."""
        result = self.db.execute(
            select(User).where(User.username == request.username)
        )
        user = result.scalar_one_or_none()

        if not user:
            raise AuthenticationError("Invalid username or password")

        if not verify_password(request.password, user.password_hash):
            raise AuthenticationError("Invalid username or password")

        if not user.is_active:
            raise AuthenticationError("User account is deactivated")

        # Update last login
        user.last_login_at = datetime.now(timezone.utc)
        self.db.flush()

        # Generate tokens
        access_token = create_access_token(user.id, user.username)
        refresh_token = create_refresh_token(user.id)

        from app.core.config import settings

        return TokenResponse(
            access_token=access_token,
            refresh_token=refresh_token,
            token_type="bearer",
            expires_in=settings.JWT_ACCESS_TOKEN_EXPIRE_MINUTES * 60,
        )

    def refresh_tokens(self, refresh_token: str) -> TokenResponse:
        """Refresh access token using refresh token."""
        payload = verify_refresh_token(refresh_token)

        if not payload:
            raise AuthenticationError("Invalid or expired refresh token")

        user_id = payload.get("sub")
        if not user_id:
            raise AuthenticationError("Invalid token payload")

        try:
            user_uuid = UUID(user_id)
        except ValueError:
            raise AuthenticationError("Invalid user ID in token")

        result = self.db.execute(select(User).where(User.id == user_uuid))
        user = result.scalar_one_or_none()

        if not user or not user.is_active:
            raise AuthenticationError("User not found or deactivated")

        # Generate new tokens
        access_token = create_access_token(user.id, user.username)
        new_refresh_token = create_refresh_token(user.id)

        from app.core.config import settings

        return TokenResponse(
            access_token=access_token,
            refresh_token=new_refresh_token,
            token_type="bearer",
            expires_in=settings.JWT_ACCESS_TOKEN_EXPIRE_MINUTES * 60,
        )

    def register_user(self, request: UserCreate) -> UserResponse:
        """Register a new user."""
        # Check if username exists
        result = self.db.execute(
            select(User).where(User.username == request.username)
        )
        existing = result.scalar_one_or_none()

        if existing:
            raise ValidationError("Username already registered")

        # Create user
        user = User(
            name=request.name,
            username=request.username,
            phone=request.phone,
            password_hash=hash_password(request.password),
            is_active=True,
        )

        self.db.add(user)
        self.db.flush()
        self.db.refresh(user)

        return UserResponse.model_validate(user)

    def get_user_by_id(self, user_id: UUID) -> User:
        """Get user by ID."""
        result = self.db.execute(select(User).where(User.id == user_id))
        user = result.scalar_one_or_none()

        if not user:
            raise NotFoundError("User", str(user_id))

        return user

    def change_password(
        self,
        user_id: UUID,
        current_password: str,
        new_password: str,
    ) -> None:
        """Change user password."""
        user = self.get_user_by_id(user_id)

        if not verify_password(current_password, user.password_hash):
            raise AuthenticationError("Current password is incorrect")

        user.password_hash = hash_password(new_password)
        self.db.flush()

    def list_all_users(self) -> list[UserWithProjectRoles]:
        """List all users with their project-role mappings (for super admin)."""
        # Get all users
        result = self.db.execute(
            select(User).order_by(User.name)
        )
        users = result.scalars().all()
        
        users_with_roles = []
        for user in users:
            # Get user's role assignments
            role_result = self.db.execute(
                select(UserRoleProject, Role, Project)
                .join(Role, UserRoleProject.role_id == Role.id)
                .join(Project, UserRoleProject.project_id == Project.id)
                .where(UserRoleProject.user_id == user.id)
                .order_by(Project.name, Role.name)
            )
            role_assignments = role_result.all()
            
            # Group by project
            project_roles_map: dict[int, ProjectRoleMapping] = {}
            for urp, role, project in role_assignments:
                if project.id not in project_roles_map:
                    project_roles_map[project.id] = ProjectRoleMapping(
                        project_id=project.id,
                        project_name=project.name,
                        roles=[],
                    )
                project_roles_map[project.id].roles.append(RoleInfo(
                    id=role.id,
                    name=role.name,
                ))
            
            users_with_roles.append(UserWithProjectRoles(
                **UserResponse.model_validate(user).model_dump(),
                project_roles=list(project_roles_map.values()),
            ))
        
        return users_with_roles

    def list_unassigned_users(self) -> list[UserWithProjectRoles]:
        """List users who have no project assignments (for super admin)."""
        from sqlalchemy import not_
        
        # Get users who have no entries in UserRoleProject
        subquery = select(UserRoleProject.user_id).where(
            UserRoleProject.user_id == User.id
        ).exists()
        
        result = self.db.execute(
            select(User)
            .where(not_(subquery))
            .order_by(User.name)
        )
        users = result.scalars().all()
        
        # These users have no project assignments, so project_roles will be empty
        users_with_roles = []
        for user in users:
            users_with_roles.append(UserWithProjectRoles(
                **UserResponse.model_validate(user).model_dump(),
                project_roles=[],
            ))
        
        return users_with_roles

    def update_user_admin(
        self,
        user_id: int,
        request: AdminUserUpdate,
    ) -> UserResponse:
        """Update user as admin (can toggle active/super admin)."""
        result = self.db.execute(select(User).where(User.id == user_id))
        user = result.scalar_one_or_none()

        if not user:
            raise NotFoundError("User", str(user_id))

        update_data = request.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(user, field, value)

        self.db.flush()
        self.db.refresh(user)

        return UserResponse.model_validate(user)

    def delete_user(self, user_id: int) -> None:
        """Delete a user (for super admin)."""
        result = self.db.execute(select(User).where(User.id == user_id))
        user = result.scalar_one_or_none()

        if not user:
            raise NotFoundError("User", str(user_id))

        self.db.delete(user)
        self.db.flush()

    def is_user_in_project(self, user_id: int, project_id: int) -> bool:
        """Check if a user belongs to a specific project."""
        result = self.db.execute(
            select(UserRoleProject)
            .where(
                UserRoleProject.user_id == user_id,
                UserRoleProject.project_id == project_id,
            )
        )
        return result.scalar_one_or_none() is not None

    def update_project_user(
        self,
        user_id: int,
        project_id: int,
        request: ProjectUserUpdate,
    ) -> UserResponse:
        """Update user within a project context (for school admin)."""
        # First verify user is in this project
        if not self.is_user_in_project(user_id, project_id):
            raise NotFoundError("User", str(user_id))

        result = self.db.execute(select(User).where(User.id == user_id))
        user = result.scalar_one_or_none()

        if not user:
            raise NotFoundError("User", str(user_id))

        update_data = request.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(user, field, value)

        self.db.flush()
        self.db.refresh(user)

        return UserResponse.model_validate(user)

    def remove_user_from_project(self, user_id: int, project_id: int) -> None:
        """Remove a user from a project (removes their role assignment)."""
        # Check if user is in project
        result = self.db.execute(
            select(UserRoleProject)
            .where(
                UserRoleProject.user_id == user_id,
                UserRoleProject.project_id == project_id,
            )
        )
        assignment = result.scalar_one_or_none()

        if not assignment:
            raise NotFoundError("User", str(user_id))

        self.db.delete(assignment)
        self.db.flush()

    def generate_user_template(self, include_role_column: bool = False) -> bytes:
        """Generate Excel template for user bulk upload."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"

        # Write headers
        headers = [col[1] for col in USER_TEMPLATE_COLUMNS]
        if include_role_column:
            headers.append("Role Name")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = cell.font.copy(bold=True)

        # Add sample row
        sample_data = ["John Doe", "johndoe", "+1234567890", "password123"]
        if include_role_column:
            sample_data.append("Staff")  # Example role name
        
        for col_idx, value in enumerate(sample_data, start=1):
            ws.cell(row=2, column=col_idx, value=value)

        # Adjust column widths
        column_widths = [25, 20, 15, 15]
        if include_role_column:
            column_widths.append(10)
        
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def bulk_upload_users(
        self,
        file_content: bytes,
        project_id: int | None = None,
        default_role_id: int | None = None,
    ) -> UserBulkUploadResult:
        """Process bulk user upload from Excel.
        
        Args:
            file_content: Excel file content
            project_id: Optional project to assign users to
            default_role_id: Optional default role ID to assign if not specified in Excel
        """
        try:
            wb = load_workbook(BytesIO(file_content), data_only=True)
            ws = wb.active
        except Exception as e:
            raise ValidationError(f"Invalid Excel file: {str(e)}")

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            raise ValidationError("No data found in Excel file")

        errors = []
        successful = 0
        created_users = []

        for row_num, row in enumerate(rows, start=2):
            try:
                # Skip empty rows
                if not any(row):
                    continue

                # Parse row data
                name = str(row[0]).strip() if row[0] else None
                username = str(row[1]).strip() if row[1] else None
                phone = str(row[2]).strip() if len(row) > 2 and row[2] else None
                password = str(row[3]).strip() if len(row) > 3 and row[3] else None
                role_name = str(row[4]).strip() if len(row) > 4 and row[4] else None

                # Validate required fields
                if not name:
                    raise ValidationError("Full Name is required", details={"column": "Full Name", "row": row_num})
                if not username:
                    raise ValidationError("Username is required", details={"column": "Username", "row": row_num})
                if not password:
                    raise ValidationError("Password is required", details={"column": "Password", "row": row_num})
                if len(password) < 8:
                    raise ValidationError("Password must be at least 8 characters", details={"column": "Password", "row": row_num})

                # Check if username exists
                result = self.db.execute(
                    select(User).where(User.username == username)
                )
                if result.scalar_one_or_none():
                    raise ValidationError(f"Username '{username}' already exists", details={"column": "Username", "row": row_num})

                # Create user
                user = User(
                    name=name,
                    username=username,
                    phone=phone,
                    password_hash=hash_password(password),
                    is_active=True,
                )
                self.db.add(user)
                self.db.flush()
                
                # Assign role if project_id is provided
                if project_id:
                    role = None
                    
                    # Try to find role by name (case-insensitive)
                    if role_name:
                        role_result = self.db.execute(
                            select(Role).where(
                                Role.project_id == project_id,
                            )
                        )
                        project_roles = role_result.scalars().all()
                        for r in project_roles:
                            if r.name.lower() == role_name.lower():
                                role = r
                                break
                        
                        if not role:
                            raise ValidationError(
                                f"Role '{role_name}' not found in this project",
                                details={"column": "Role Name", "row": row_num}
                            )
                    elif default_role_id:
                        # Use default role if no role name provided
                        role_result = self.db.execute(
                            select(Role).where(
                                Role.id == default_role_id,
                                Role.project_id == project_id,
                            )
                        )
                        role = role_result.scalar_one_or_none()
                    
                    if role:
                        user_role = UserRoleProject(
                            user_id=user.id,
                            role_id=role.id,
                            project_id=project_id,
                        )
                        self.db.add(user_role)
                
                created_users.append(user)
                successful += 1

            except ValidationError as e:
                errors.append({
                    "row": row_num,
                    "column": e.details.get("column") if e.details else None,
                    "message": e.message,
                })
            except Exception as e:
                errors.append({
                    "row": row_num,
                    "message": str(e),
                })

        self.db.flush()

        total = len([r for r in rows if any(r)])
        message = f"Created {successful} of {total} users."
        if errors:
            message += f" {len(errors)} rows failed."

        return UserBulkUploadResult(
            total_rows=total,
            successful_rows=successful,
            failed_rows=len(errors),
            errors=errors,
            message=message,
        )

    def list_project_users(self, project_id: int) -> list[UserWithProjectRoles]:
        """List all users in a specific project with their roles."""
        # Get all users in this project
        result = self.db.execute(
            select(User)
            .join(UserRoleProject, User.id == UserRoleProject.user_id)
            .where(UserRoleProject.project_id == project_id)
            .distinct()
            .order_by(User.name)
        )
        users = result.scalars().all()
        
        users_with_roles = []
        for user in users:
            # Get user's role assignments in this project
            role_result = self.db.execute(
                select(UserRoleProject, Role)
                .join(Role, UserRoleProject.role_id == Role.id)
                .where(
                    UserRoleProject.user_id == user.id,
                    UserRoleProject.project_id == project_id,
                )
                .order_by(Role.name)
            )
            role_assignments = role_result.all()
            
            # Get project info
            project_result = self.db.execute(
                select(Project).where(Project.id == project_id)
            )
            project = project_result.scalar_one_or_none()
            
            project_roles_map: dict[int, ProjectRoleMapping] = {}
            if project:
                project_roles_map[project_id] = ProjectRoleMapping(
                    project_id=project_id,
                    project_name=project.name,
                    roles=[],
                )
                for urp, role in role_assignments:
                    project_roles_map[project_id].roles.append(RoleInfo(
                        id=role.id,
                        name=role.name,
                    ))
            
            users_with_roles.append(UserWithProjectRoles(
                **UserResponse.model_validate(user).model_dump(),
                project_roles=list(project_roles_map.values()),
            ))
        
        return users_with_roles
