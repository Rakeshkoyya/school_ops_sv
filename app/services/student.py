"""Student management service."""

from io import BytesIO
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import NotFoundError, ValidationError
from app.models.student import Student
from app.schemas.student import (
    StudentCreate,
    StudentResponse,
    StudentUpdate,
    StudentBulkUploadResult,
    StudentFilter,
    PaginatedStudentResponse,
)


# Excel template columns for student upload
STUDENT_TEMPLATE_COLUMNS = [
    ("student_name", "Student Name", True),
    ("class_name", "Class", True),
    ("section", "Section", False),
    ("parent_name", "Parent Name", False),
    ("parent_phone_no", "Parent Phone", False),
]


class StudentService:
    """Student management service."""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def create_student(
        self,
        project_id: int,
        request: StudentCreate,
    ) -> StudentResponse:
        """Create a new student."""
        student = Student(
            project_id=project_id,
            student_name=request.student_name,
            class_name=request.class_name,
            section=request.section,
            parent_name=request.parent_name,
            parent_phone_no=request.parent_phone_no,
        )
        self.db.add(student)
        await self.db.flush()
        await self.db.refresh(student)
        return StudentResponse.model_validate(student)

    async def get_student(self, project_id: int, student_id: int) -> Student:
        """Get student by ID."""
        result = await self.db.execute(
            select(Student).where(
                Student.id == student_id,
                Student.project_id == project_id,
            )
        )
        student = result.scalar_one_or_none()
        if not student:
            raise NotFoundError("Student", str(student_id))
        return student

    async def update_student(
        self,
        project_id: int,
        student_id: int,
        request: StudentUpdate,
    ) -> StudentResponse:
        """Update a student."""
        student = await self.get_student(project_id, student_id)
        update_data = request.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(student, field, value)
        await self.db.flush()
        await self.db.refresh(student)
        return StudentResponse.model_validate(student)

    async def delete_student(self, project_id: int, student_id: int) -> None:
        """Delete a student."""
        student = await self.get_student(project_id, student_id)
        await self.db.delete(student)
        await self.db.flush()

    async def list_students(
        self,
        project_id: int,
        filters: StudentFilter | None = None,
        page: int = 1,
        page_size: int = 50,
    ) -> PaginatedStudentResponse:
        """List students with filtering and pagination."""
        query = select(Student).where(Student.project_id == project_id)

        if filters:
            if filters.class_name:
                query = query.where(Student.class_name == filters.class_name)
            if filters.section:
                query = query.where(Student.section == filters.section)
            if filters.search:
                search_term = f"%{filters.search}%"
                query = query.where(
                    or_(
                        Student.student_name.ilike(search_term),
                        Student.parent_name.ilike(search_term),
                    )
                )

        # Get total count
        count_query = select(func.count()).select_from(query.subquery())
        total_result = await self.db.execute(count_query)
        total = total_result.scalar() or 0

        # Apply pagination
        offset = (page - 1) * page_size
        query = query.order_by(Student.class_name, Student.section, Student.student_name)
        query = query.offset(offset).limit(page_size)

        result = await self.db.execute(query)
        students = result.scalars().all()

        return PaginatedStudentResponse(
            items=[StudentResponse.model_validate(s) for s in students],
            total=total,
            page=page,
            page_size=page_size,
            total_pages=(total + page_size - 1) // page_size,
        )

    def generate_template(self) -> bytes:
        """Generate Excel template for student bulk upload."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        # Write headers
        headers = [col[1] for col in STUDENT_TEMPLATE_COLUMNS]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = cell.font.copy(bold=True)

        # Add sample row
        sample_data = ["John Doe", "10", "A", "Mr. Doe", "+1234567890"]
        for col_idx, value in enumerate(sample_data, start=1):
            ws.cell(row=2, column=col_idx, value=value)

        # Adjust column widths
        column_widths = [25, 10, 10, 25, 20]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def bulk_upload(
        self,
        project_id: int,
        file_content: bytes,
    ) -> StudentBulkUploadResult:
        """Process bulk student upload from Excel."""
        try:
            wb = load_workbook(BytesIO(file_content), data_only=True)
            ws = wb.active
        except Exception as e:
            raise ValidationError(f"Invalid Excel file: {str(e)}")

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            raise ValidationError("No data found in Excel file")

        errors = []
        successful = 0

        for row_num, row in enumerate(rows, start=2):
            try:
                # Skip empty rows
                if not any(row):
                    continue

                # Parse row data
                student_name = str(row[0]).strip() if row[0] else None
                class_name = str(row[1]).strip() if row[1] else None
                section = str(row[2]).strip() if len(row) > 2 and row[2] else None
                parent_name = str(row[3]).strip() if len(row) > 3 and row[3] else None
                parent_phone = str(row[4]).strip() if len(row) > 4 and row[4] else None

                # Validate required fields
                if not student_name:
                    raise ValidationError("Student Name is required", details={"column": "Student Name", "row": row_num})
                if not class_name:
                    raise ValidationError("Class is required", details={"column": "Class", "row": row_num})

                # Create student
                student = Student(
                    project_id=project_id,
                    student_name=student_name,
                    class_name=class_name,
                    section=section,
                    parent_name=parent_name,
                    parent_phone_no=parent_phone,
                )
                self.db.add(student)
                successful += 1

            except ValidationError as e:
                errors.append({
                    "row": row_num,
                    "column": e.details.get("column") if e.details else None,
                    "message": e.message,
                })
            except Exception as e:
                errors.append({
                    "row": row_num,
                    "message": str(e),
                })

        await self.db.flush()

        total = len([r for r in rows if any(r)])
        message = f"Uploaded {successful} of {total} students."
        if errors:
            message += f" {len(errors)} rows failed."

        return StudentBulkUploadResult(
            total_rows=total,
            successful_rows=successful,
            failed_rows=len(errors),
            errors=errors,
            message=message,
        )

    async def get_class_sections(self, project_id: int) -> list[dict]:
        """Get distinct class-section combinations for a project."""
        result = await self.db.execute(
            select(Student.class_name, Student.section)
            .where(Student.project_id == project_id)
            .distinct()
            .order_by(Student.class_name, Student.section)
        )
        rows = result.all()
        return [{"class_name": r[0], "section": r[1]} for r in rows]
