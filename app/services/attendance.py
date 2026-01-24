"""Attendance service for CRUD and bulk operations."""

import calendar
import logging
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

# Setup debug logger
logger = logging.getLogger(__name__)
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select, and_, or_, Integer
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import NotFoundError, ValidationError
from app.models.attendance import AttendanceRecord, AttendanceStatus
from app.models.student import Student
from app.schemas.attendance import (
    AttendanceByClassResponse,
    AttendanceFilter,
    AttendanceRecordCreate,
    AttendanceRecordResponse,
    AttendanceRecordUpdate,
    AttendanceSummary,
    AttendanceUploadError,
    AttendanceUploadResult,
    BulkAttendanceCreate,
    BulkAttendanceResponse,
    SingleAttendanceInput,
)


class AttendanceService:
    """Attendance management service."""

    def __init__(self, db: AsyncSession):
        self.db = db

    def _record_to_response(self, record: AttendanceRecord) -> dict:
        """Convert AttendanceRecord to response dict."""
        return {
            "id": record.id,
            "project_id": record.project_id,
            "student_id": record.student_id,
            "student_name": record.student_name,
            "class_name": record.class_name,
            "section": record.section,
            "attendance_date": record.attendance_date,
            "status": record.status,
            "remarks": record.remarks,
            "upload_id": record.upload_id,
            "created_at": record.created_at,
            "updated_at": record.updated_at,
        }

    async def create_record(
        self,
        project_id: int,
        request: AttendanceRecordCreate,
    ) -> AttendanceRecordResponse:
        """Create a single attendance record."""
        # Verify student exists and belongs to project
        student = await self._get_student(project_id, request.student_id)
        
        # Check for existing record
        existing = await self._get_existing_record(
            project_id, request.student_id, request.attendance_date
        )
        if existing:
            raise ValidationError(
                f"Attendance already exists for {student.student_name} on {request.attendance_date}"
            )

        record = AttendanceRecord(
            project_id=project_id,
            student_id=request.student_id,
            attendance_date=request.attendance_date,
            status=request.status,
            remarks=request.remarks,
        )
        self.db.add(record)
        await self.db.flush()
        await self.db.refresh(record)

        return AttendanceRecordResponse.model_validate(self._record_to_response(record))

    async def _get_student(self, project_id: int, student_id: int) -> Student:
        """Get student by ID, validating project membership."""
        result = await self.db.execute(
            select(Student).where(
                Student.id == student_id,
                Student.project_id == project_id,
            )
        )
        student = result.scalar_one_or_none()
        if not student:
            raise NotFoundError("Student", str(student_id))
        return student

    async def _get_existing_record(
        self, project_id: int, student_id: int, attendance_date: date
    ) -> AttendanceRecord | None:
        """Check for existing attendance record."""
        result = await self.db.execute(
            select(AttendanceRecord).where(
                AttendanceRecord.project_id == project_id,
                AttendanceRecord.student_id == student_id,
                AttendanceRecord.attendance_date == attendance_date,
            )
        )
        return result.scalar_one_or_none()

    async def get_record(
        self,
        record_id: int,
        project_id: int,
    ) -> AttendanceRecord:
        """Get attendance record by ID."""
        result = await self.db.execute(
            select(AttendanceRecord).where(
                AttendanceRecord.id == record_id,
                AttendanceRecord.project_id == project_id,
            )
        )
        record = result.scalar_one_or_none()
        if not record:
            raise NotFoundError("Attendance record", str(record_id))
        return record

    async def list_records(
        self,
        project_id: int,
        filters: AttendanceFilter | None = None,
        page: int = 1,
        page_size: int = 50,
    ) -> tuple[list[AttendanceRecordResponse], int]:
        """List attendance records with filtering."""
        query = select(AttendanceRecord).where(
            AttendanceRecord.project_id == project_id
        )

        if filters:
            if filters.student_id:
                query = query.where(AttendanceRecord.student_id == filters.student_id)
            if filters.status:
                query = query.where(AttendanceRecord.status == filters.status)
            if filters.date_from:
                query = query.where(AttendanceRecord.attendance_date >= filters.date_from)
            if filters.date_to:
                query = query.where(AttendanceRecord.attendance_date <= filters.date_to)
            if filters.class_section:
                # Parse class_section like "3-A" into class_name and section
                class_name, section = self._parse_class_section(filters.class_section)
                # Join with student to filter by class
                query = query.join(Student).where(
                    Student.class_name == class_name
                )
                if section:
                    query = query.where(Student.section == section)
            elif filters.class_name:
                query = query.join(Student).where(Student.class_name == filters.class_name)
                if filters.section:
                    query = query.where(Student.section == filters.section)

        # Count total
        count_result = await self.db.execute(
            select(func.count()).select_from(query.subquery())
        )
        total = count_result.scalar() or 0

        # Apply pagination and ordering
        query = (
            query
            .order_by(AttendanceRecord.attendance_date.desc(), AttendanceRecord.id)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )

        result = await self.db.execute(query)
        records = result.scalars().all()

        return [
            AttendanceRecordResponse.model_validate(self._record_to_response(r))
            for r in records
        ], total

    async def update_record(
        self,
        record_id: int,
        project_id: int,
        request: AttendanceRecordUpdate,
    ) -> AttendanceRecordResponse:
        """Update an attendance record."""
        record = await self.get_record(record_id, project_id)

        update_data = request.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(record, field, value)

        await self.db.flush()
        await self.db.refresh(record)

        return AttendanceRecordResponse.model_validate(self._record_to_response(record))

    async def delete_record(
        self,
        record_id: int,
        project_id: int,
    ) -> None:
        """Delete an attendance record."""
        record = await self.get_record(record_id, project_id)
        await self.db.delete(record)
        await self.db.flush()

    async def get_summary(
        self,
        project_id: int,
        date_from: date,
        date_to: date,
        student_id: int | None = None,
        class_section: str | None = None,
    ) -> AttendanceSummary:
        """Get attendance summary for a date range."""
        query = select(
            func.count().label("total"),
            func.sum(func.cast(AttendanceRecord.status == AttendanceStatus.PRESENT, Integer)).label("present"),
            func.sum(func.cast(AttendanceRecord.status == AttendanceStatus.ABSENT, Integer)).label("absent"),
            func.sum(func.cast(AttendanceRecord.status == AttendanceStatus.LATE, Integer)).label("late"),
            func.sum(func.cast(AttendanceRecord.status == AttendanceStatus.EXCUSED, Integer)).label("excused"),
        ).where(
            AttendanceRecord.project_id == project_id,
            AttendanceRecord.attendance_date >= date_from,
            AttendanceRecord.attendance_date <= date_to,
        )

        if student_id:
            query = query.where(AttendanceRecord.student_id == student_id)

        if class_section:
            class_name, section = self._parse_class_section(class_section)
            query = query.join(Student).where(Student.class_name == class_name)
            if section:
                query = query.where(Student.section == section)

        result = await self.db.execute(query)
        row = result.one()

        return AttendanceSummary(
            total_records=row.total or 0,
            present_count=row.present or 0,
            absent_count=row.absent or 0,
            late_count=row.late or 0,
            excused_count=row.excused or 0,
            date_from=date_from,
            date_to=date_to,
        )

    # ==========================================
    # Bulk Operations
    # ==========================================

    async def bulk_create_or_update(
        self,
        project_id: int,
        request: BulkAttendanceCreate,
    ) -> BulkAttendanceResponse:
        """Create or update attendance records in bulk for a class on a specific date."""
        errors = []
        successful = 0
        failed = 0

        # Parse class section
        class_name, section = self._parse_class_section(request.class_section)

        # Get all students for the class
        students = await self._get_students_by_class(project_id, class_name, section)
        student_ids = {s.id for s in students}

        # Validate all student IDs before any DB operations
        for record in request.records:
            if record.student_id not in student_ids:
                errors.append({
                    "student_id": record.student_id,
                    "message": f"Student ID {record.student_id} not found in class {request.class_section}",
                })
                failed += 1

        if errors:
            return BulkAttendanceResponse(
                total_records=len(request.records),
                successful=0,
                failed=failed,
                errors=errors,
                message="Validation failed. No records were saved.",
            )

        # All validations passed, proceed with DB operations
        for record in request.records:
            try:
                existing = await self._get_existing_record(
                    project_id, record.student_id, request.attendance_date
                )
                
                if existing:
                    # Update existing record
                    existing.status = record.status
                    existing.remarks = record.remarks
                else:
                    # Create new record
                    new_record = AttendanceRecord(
                        project_id=project_id,
                        student_id=record.student_id,
                        attendance_date=request.attendance_date,
                        status=record.status,
                        remarks=record.remarks,
                    )
                    self.db.add(new_record)
                
                successful += 1
            except Exception as e:
                errors.append({
                    "student_id": record.student_id,
                    "message": str(e),
                })
                failed += 1

        await self.db.flush()

        return BulkAttendanceResponse(
            total_records=len(request.records),
            successful=successful,
            failed=failed,
            errors=errors,
            message=f"Successfully saved {successful} attendance records.",
        )

    async def get_attendance_by_class_date(
        self,
        project_id: int,
        class_section: str,
        attendance_date: date,
    ) -> AttendanceByClassResponse:
        """Get all student attendance for a class on a specific date."""
        class_name, section = self._parse_class_section(class_section)

        # Get all students in the class
        students = await self._get_students_by_class(project_id, class_name, section)

        # Get existing attendance records for the date
        result = await self.db.execute(
            select(AttendanceRecord).where(
                AttendanceRecord.project_id == project_id,
                AttendanceRecord.attendance_date == attendance_date,
                AttendanceRecord.student_id.in_([s.id for s in students]),
            )
        )
        records = {r.student_id: r for r in result.scalars().all()}

        # Build response with all students
        student_data = []
        present_count = 0
        absent_count = 0
        late_count = 0
        excused_count = 0

        for student in students:
            record = records.get(student.id)
            status = record.status if record else None
            
            if status == AttendanceStatus.PRESENT:
                present_count += 1
            elif status == AttendanceStatus.ABSENT:
                absent_count += 1
            elif status == AttendanceStatus.LATE:
                late_count += 1
            elif status == AttendanceStatus.EXCUSED:
                excused_count += 1

            student_data.append({
                "student_id": student.id,
                "student_name": student.student_name,
                "class_name": student.class_name,
                "section": student.section,
                "status": status.value if status else None,
                "remarks": record.remarks if record else None,
                "record_id": record.id if record else None,
            })

        return AttendanceByClassResponse(
            class_section=class_section,
            attendance_date=attendance_date,
            students=student_data,
            total_students=len(students),
            present_count=present_count,
            absent_count=absent_count,
            late_count=late_count,
            excused_count=excused_count,
        )

    # ==========================================
    # Template Generation
    # ==========================================

    async def generate_template(
        self,
        project_id: int,
        class_section: str | None = None,
        month: int | None = None,
        year: int | None = None,
    ) -> bytes:
        """Generate Excel template for attendance upload.
        
        Template format:
        Student Name | Grade | Week 1 (Mon-Sat) | Week 2 (Mon-Sat) | ...
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Default to current month/year
        today = date.today()
        month = month or today.month
        year = year or today.year

        # Get month details
        month_name = calendar.month_abbr[month]
        _, days_in_month = calendar.monthrange(year, month)

        # Generate week headers
        headers = ["Student Name", "Grade"]
        date_columns = []  # Track (column_index, date) for each day

        current_date = date(year, month, 1)
        week_num = 1
        week_start_col = 3

        while current_date.month == month:
            # Get day of week (0=Monday, 6=Sunday)
            day_of_week = current_date.weekday()
            
            # Skip Sundays (weekday 6)
            if day_of_week == 6:
                current_date += timedelta(days=1)
                continue

            # Start new week on Monday
            if day_of_week == 0:
                week_start_col = len(headers) + 1

            day_name = calendar.day_abbr[day_of_week][:3]
            headers.append(day_name)
            date_columns.append((len(headers), current_date))
            
            current_date += timedelta(days=1)

        # Styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        # Write week header row (Row 1) - merged cells for weeks
        current_date = date(year, month, 1)
        col_idx = 3
        week_ranges = []
        week_start = col_idx

        while current_date.month == month:
            day_of_week = current_date.weekday()
            
            if day_of_week == 6:  # Sunday
                current_date += timedelta(days=1)
                continue

            if day_of_week == 0 and col_idx > 3:  # New week starting
                week_ranges.append((week_start, col_idx - 1, week_num))
                week_num += 1
                week_start = col_idx

            col_idx += 1
            current_date += timedelta(days=1)

        # Add last week
        if week_start <= len(headers):
            week_ranges.append((week_start, len(headers), week_num))

        # Merge and format week headers
        for start_col, end_col, week in week_ranges:
            if start_col <= end_col:
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                cell = ws.cell(row=1, column=start_col)
                cell.value = f"{year} {month_name} Week {week}"
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

        # Write day headers (Row 2)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_align
            if col_idx <= 2:
                cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Add students if class_section is provided
        if class_section:
            class_name, section = self._parse_class_section(class_section)
            students = await self._get_students_by_class(project_id, class_name, section)

            for row_idx, student in enumerate(students, start=3):
                ws.cell(row=row_idx, column=1, value=student.student_name).border = thin_border
                grade_str = f"{student.class_name}-{student.section}" if student.section else student.class_name
                ws.cell(row=row_idx, column=2, value=grade_str).border = thin_border
                
                # Add empty cells with borders for attendance
                for col in range(3, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col, value="P")  # Default to Present
                    cell.border = thin_border
                    cell.alignment = center_align

        # Adjust column widths
        ws.column_dimensions['A'].width = 25  # Student Name
        ws.column_dimensions['B'].width = 10  # Grade
        for col in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 5

        # Add instructions sheet
        instructions_ws = wb.create_sheet("Instructions")
        instructions = [
            ("Attendance Template Instructions", ""),
            ("", ""),
            ("Status Values:", "Use the following values for attendance:"),
            ("P", "Present"),
            ("A", "Absent"),
            ("L", "Late"),
            ("E", "Excused"),
            ("", ""),
            ("Notes:", ""),
            ("- Each column represents a day (Mon-Sat)", ""),
            ("- Sundays are excluded", ""),
            ("- Leave cell empty if no attendance data", ""),
            ("- Student names must match exactly with system records", ""),
        ]
        for row_idx, (col1, col2) in enumerate(instructions, start=1):
            instructions_ws.cell(row=row_idx, column=1, value=col1)
            instructions_ws.cell(row=row_idx, column=2, value=col2)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # ==========================================
    # Excel Upload Processing
    # ==========================================

    async def process_excel_upload(
        self,
        project_id: int,
        file_content: bytes,
        upload_id: int | None = None,
    ) -> AttendanceUploadResult:
        """Process attendance Excel upload with validation."""
        logger.info(f"[UPLOAD] Starting - project_id={project_id}, file_size={len(file_content)} bytes")
        
        try:
            wb = load_workbook(BytesIO(file_content), data_only=True)
            ws = wb.active
        except Exception as e:
            logger.error(f"[UPLOAD] Failed to load Excel: {str(e)}")
            raise ValidationError(f"Invalid Excel file: {str(e)}")

        errors: list[AttendanceUploadError] = []
        successful_rows = 0
        failed_rows = 0
        skipped_rows = 0

        # Parse headers to get dates
        week_headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        day_headers = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

        # Build date mapping for each column
        date_columns = self._parse_template_headers(week_headers, day_headers)
        
        # Log detected dates clearly
        detected_dates = sorted(set(date_columns.values()))
        logger.info(f"[UPLOAD] Detected {len(detected_dates)} dates from Excel: {detected_dates}")

        if not date_columns:
            logger.error("[UPLOAD] No dates could be parsed from headers")
            raise ValidationError("Could not parse date headers from template")

        # Get all students in project indexed by name (case-insensitive)
        students_result = await self.db.execute(
            select(Student).where(Student.project_id == project_id)
        )
        students_by_name: dict[str, Student] = {}
        for s in students_result.scalars().all():
            key = s.student_name.lower().strip()
            students_by_name[key] = s
        logger.info(f"[UPLOAD] Found {len(students_by_name)} students in database")

        # Track counts per day
        counts_per_day: dict[date, dict[str, int]] = {d: {"present": 0, "absent": 0, "late": 0, "excused": 0} for d in detected_dates}
        
        # First pass: Validate all data
        rows_to_process = []
        students_not_found = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not any(row):
                continue

            student_name = str(row[0]).strip() if row[0] else None
            grade = str(row[1]).strip() if row[1] else None

            if not student_name:
                skipped_rows += 1
                continue

            student_key = student_name.lower()
            student = students_by_name.get(student_key)

            if not student:
                students_not_found.append(student_name)
                errors.append(AttendanceUploadError(
                    row=row_num,
                    student_name=student_name,
                    message=f"Student '{student_name}' not found in system",
                ))
                failed_rows += 1
                continue

            # Verify grade matches
            expected_grade = f"{student.class_name}-{student.section}" if student.section else student.class_name
            if grade and grade != expected_grade:
                errors.append(AttendanceUploadError(
                    row=row_num,
                    student_name=student_name,
                    column="Grade",
                    message=f"Grade mismatch: expected '{expected_grade}', got '{grade}'",
                ))
                failed_rows += 1
                continue

            # Parse attendance for each date column
            row_valid = True
            attendance_entries = []

            for col_idx, att_date in date_columns.items():
                if col_idx - 1 >= len(row):
                    continue
                    
                value = row[col_idx - 1]
                if not value or str(value).strip() == "":
                    continue

                value_str = str(value).strip().upper()
                try:
                    status = AttendanceStatus.from_string(value_str)
                    attendance_entries.append((att_date, status))
                    # Track count per day
                    if att_date in counts_per_day:
                        counts_per_day[att_date][status.value] += 1
                except ValueError:
                    errors.append(AttendanceUploadError(
                        row=row_num,
                        student_name=student_name,
                        column=day_headers[col_idx - 1] if col_idx - 1 < len(day_headers) else f"Column {col_idx}",
                        message=f"Invalid status value: '{value}'",
                    ))
                    row_valid = False

            if row_valid and attendance_entries:
                rows_to_process.append((student, attendance_entries))

        # Log students not found summary
        if students_not_found:
            logger.warning(f"[UPLOAD] Students not found ({len(students_not_found)}): {students_not_found[:10]}{'...' if len(students_not_found) > 10 else ''}")

        # Check for critical errors before processing
        if failed_rows > 0 and not rows_to_process:
            return AttendanceUploadResult(
                total_rows=failed_rows + skipped_rows,
                successful_rows=0,
                failed_rows=failed_rows,
                skipped_rows=skipped_rows,
                errors=errors,
                message="All rows failed validation. No records were saved.",
            )

        # Second pass: Apply validated data to database
        inserted = 0
        updated = 0
        for student, entries in rows_to_process:
            try:
                for att_date, status in entries:
                    existing = await self._get_existing_record(
                        project_id, student.id, att_date
                    )
                    
                    if existing:
                        existing.status = status
                        if upload_id:
                            existing.upload_id = upload_id
                        updated += 1
                    else:
                        record = AttendanceRecord(
                            project_id=project_id,
                            student_id=student.id,
                            attendance_date=att_date,
                            status=status,
                            upload_id=upload_id,
                        )
                        self.db.add(record)
                        inserted += 1
                
                successful_rows += 1
            except Exception as e:
                errors.append(AttendanceUploadError(
                    row=0,
                    student_name=student.student_name,
                    message=f"Database error: {str(e)}",
                ))
                failed_rows += 1

        await self.db.flush()

        # Log summary with counts per day
        total = successful_rows + failed_rows + skipped_rows
        logger.info(f"[UPLOAD] === SUMMARY ===")
        logger.info(f"[UPLOAD] Students: {successful_rows} OK, {failed_rows} failed, {skipped_rows} skipped")
        logger.info(f"[UPLOAD] Records: {inserted} inserted, {updated} updated")
        logger.info(f"[UPLOAD] Attendance per day:")
        for d in sorted(counts_per_day.keys()):
            counts = counts_per_day[d]
            total_day = sum(counts.values())
            if total_day > 0:
                logger.info(f"[UPLOAD]   {d}: P={counts['present']}, A={counts['absent']}, L={counts['late']}, E={counts['excused']} (total: {total_day})")
        
        return AttendanceUploadResult(
            total_rows=total,
            successful_rows=successful_rows,
            failed_rows=failed_rows,
            skipped_rows=skipped_rows,
            errors=errors,
            message=f"Processed {successful_rows} student records successfully.",
        )

    def _parse_template_headers(
        self, week_headers: tuple, day_headers: tuple
    ) -> dict[int, date]:
        """Parse template headers to extract date for each column.
        
        Template format matches generate_template():
        - Week 1 starts from the first day of the month (not first Monday)
        - Only Sundays are skipped
        - New week starts on Monday
        
        Example: January 2026
        - Jan 1 = Thursday → Week 1 has Thu(1), Fri(2), Sat(3)
        - Jan 5 = Monday → Week 2 has Mon(5), Tue(6), Wed(7), Thu(8), Fri(9), Sat(10)
        
        Note: Excel merged cells only have a value in the first cell of the merge.
        We need to first map all columns to their week headers, then parse dates.
        """
        date_columns = {}
        
        # First pass: Build a list of all week headers, expanding merged cells
        # Week headers are only in columns >= 3 (index 2+)
        expanded_week_headers = list(week_headers)
        last_week_header = None
        
        for col_idx in range(len(expanded_week_headers)):
            if col_idx < 2:
                # Skip "Student Name" and "Grade" columns
                continue
            
            current_header = expanded_week_headers[col_idx]
            if current_header and isinstance(current_header, str) and "Week" in current_header:
                # This is a valid week header
                last_week_header = current_header
            elif last_week_header:
                # This column is part of a merged cell, use the last seen week header
                expanded_week_headers[col_idx] = last_week_header
        
        # Second pass: Parse dates for each column
        for col_idx in range(len(expanded_week_headers)):
            if col_idx < 2:
                continue
                
            week_header = expanded_week_headers[col_idx]
            if not week_header or not isinstance(week_header, str):
                continue
            
            # Parse "2026 Jan Week 1" format
            parts = week_header.split()
            if len(parts) < 4:
                continue
                
            try:
                current_year = int(parts[0])
                month_str = parts[1]
                # Convert month name to number
                month_abbrs = {name: num for num, name in enumerate(calendar.month_abbr) if name}
                current_month = month_abbrs.get(month_str[:3].title())
                current_week_num = int(parts[-1])
                
                if not current_month:
                    logger.warning(f"[PARSE HEADER] Could not parse month from '{month_str}'")
                    continue
                    
            except (ValueError, IndexError) as e:
                logger.warning(f"[PARSE HEADER] Failed to parse '{week_header}': {e}")
                continue

            # Map day headers to dates using the same logic as generate_template
            day_header = day_headers[col_idx] if col_idx < len(day_headers) else None
            if day_header:
                day_name = str(day_header).strip()[:3].title()
                # Map day name to weekday number (Mon=0, Tue=1, ..., Sat=5)
                day_name_to_weekday = {"Mon": 0, "Tue": 1, "Wed": 2, "Thu": 3, "Fri": 4, "Sat": 5}
                
                if day_name in day_name_to_weekday:
                    target_weekday = day_name_to_weekday[day_name]
                    
                    # Calculate the target date by simulating the template generation
                    target_date = self._calculate_date_for_week_day(
                        current_year, current_month, current_week_num, target_weekday
                    )
                    
                    if target_date and target_date.month == current_month:
                        date_columns[col_idx + 1] = target_date  # +1 for 1-based indexing

        return date_columns

    def _calculate_date_for_week_day(
        self, year: int, month: int, week_num: int, target_weekday: int
    ) -> date | None:
        """Calculate the date for a specific weekday in a specific week of the month.
        
        This mirrors the generate_template() logic:
        - Week 1 starts from the first WORKING day of the month (not first calendar day)
        - New week starts on Monday (but not on the very first working day)
        - Sundays are skipped (not counted)
        
        Args:
            year: Year
            month: Month (1-12)
            week_num: Week number (1-based, as used in template)
            target_weekday: Target weekday (0=Mon, 1=Tue, ..., 5=Sat)
            
        Returns:
            The date, or None if not valid
        """
        first_day = date(year, month, 1)
        
        # Simulate walking through the month like generate_template does
        current_date = first_day
        current_week = 1
        is_first_working_day = True  # Track if this is the first working day we've seen
        
        while current_date.month == month:
            day_of_week = current_date.weekday()
            
            # Skip Sundays
            if day_of_week == 6:
                current_date += timedelta(days=1)
                continue
            
            # Check if new week starts (Monday, but NOT on the first working day)
            # This matches the template logic: col_idx > 3 (i.e., not the first day column)
            if day_of_week == 0 and not is_first_working_day:
                current_week += 1
            
            # Check if this is the target week and day
            if current_week == week_num and day_of_week == target_weekday:
                return current_date
            
            # If we've passed the target week, no point continuing
            if current_week > week_num:
                return None
            
            # After processing the first working day, mark it as done
            is_first_working_day = False
            current_date += timedelta(days=1)
        
        return None

    # ==========================================
    # Helper Methods
    # ==========================================

    def _parse_class_section(self, class_section: str) -> tuple[str, str | None]:
        """Parse class-section string like '3-A' into (class_name, section)."""
        if "-" in class_section:
            parts = class_section.rsplit("-", 1)
            return parts[0], parts[1]
        return class_section, None

    async def _get_students_by_class(
        self, project_id: int, class_name: str, section: str | None = None
    ) -> list[Student]:
        """Get all students in a class, optionally filtered by section."""
        query = select(Student).where(
            Student.project_id == project_id,
            Student.class_name == class_name,
        )
        if section:
            query = query.where(Student.section == section)
        
        query = query.order_by(Student.student_name)
        result = await self.db.execute(query)
        return list(result.scalars().all())

    async def get_class_sections(self, project_id: int) -> list[dict]:
        """Get distinct class-section combinations for a project."""
        result = await self.db.execute(
            select(Student.class_name, Student.section, func.count(Student.id).label("count"))
            .where(Student.project_id == project_id)
            .group_by(Student.class_name, Student.section)
            .order_by(Student.class_name, Student.section)
        )
        rows = result.all()
        return [
            {
                "class_name": r[0],
                "section": r[1],
                "class_section": f"{r[0]}-{r[1]}" if r[1] else r[0],
                "student_count": r[2],
            }
            for r in rows
        ]
