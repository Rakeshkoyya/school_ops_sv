"""Exam service for CRUD and bulk operations."""

import calendar
import logging
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.exceptions import NotFoundError, ValidationError
from app.models.exam import ExamRecord
from app.models.student import Student
from app.schemas.exam import (
    BulkExamCreate,
    BulkExamResponse,
    ExamByClassResponse,
    ExamFilter,
    ExamRecordCreate,
    ExamRecordResponse,
    ExamRecordUpdate,
    ExamSummary,
    ExamUploadError,
    ExamUploadResult,
    SingleExamInput,
    StudentExamEntry,
    SUBJECTS,
)

logger = logging.getLogger(__name__)


class ExamService:
    """Exam record management service."""

    def __init__(self, db: Session):
        self.db = db

    def _record_to_response(self, record: ExamRecord) -> dict:
        """Convert ExamRecord to response dict."""
        return {
            "id": record.id,
            "project_id": record.project_id,
            "student_id": record.student_id,
            "student_name": record.student.student_name if record.student else "",
            "class_name": record.student.class_name if record.student else "",
            "section": record.student.section if record.student else None,
            "exam_name": record.exam_name,
            "subject": record.subject,
            "exam_date": record.exam_date,
            "max_marks": record.max_marks,
            "marks_obtained": record.marks_obtained,
            "grade": record.grade,
            "remarks": record.remarks,
            "upload_id": record.upload_id,
            "created_at": record.created_at,
            "updated_at": record.updated_at,
        }

    def create_record(
        self,
        project_id: int,
        request: ExamRecordCreate,
    ) -> ExamRecordResponse:
        """Create a single exam record."""
        # Validate marks
        if request.marks_obtained > request.max_marks:
            raise ValidationError(
                f"marks_obtained ({request.marks_obtained}) exceeds max_marks ({request.max_marks})"
            )

        # Verify student exists and belongs to project
        student = self._get_student(project_id, request.student_id)

        # Check for existing record (same student, exam, subject, AND date)
        existing = self._get_existing_record(
            project_id, request.student_id, request.exam_name, request.subject, request.exam_date
        )
        if existing:
            raise ValidationError(
                f"Exam record already exists for {student.student_name} - {request.exam_name} - {request.subject} on {request.exam_date}"
            )

        # Calculate grade if not provided
        grade = request.grade
        if not grade:
            grade = self._calculate_grade(request.marks_obtained, request.max_marks)

        record = ExamRecord(
            project_id=project_id,
            student_id=request.student_id,
            exam_name=request.exam_name,
            subject=request.subject,
            exam_date=request.exam_date,
            max_marks=request.max_marks,
            marks_obtained=request.marks_obtained,
            grade=grade,
            remarks=request.remarks,
        )
        self.db.add(record)
        self.db.flush()
        self.db.refresh(record)

        return ExamRecordResponse.model_validate(self._record_to_response(record))

    def _get_student(self, project_id: int, student_id: int) -> Student:
        """Get student by ID, validating project membership."""
        result = self.db.execute(
            select(Student).where(
                Student.id == student_id,
                Student.project_id == project_id,
            )
        )
        student = result.scalar_one_or_none()
        if not student:
            raise NotFoundError("Student", str(student_id))
        return student

    def _get_existing_record(
        self, project_id: int, student_id: int, exam_name: str, subject: str, exam_date: date | None = None
    ) -> ExamRecord | None:
        """Check for existing exam record.
        
        Uniqueness is determined by: project_id, student_id, exam_name, subject, AND exam_date.
        This allows the same exam type to be recorded on different dates.
        """
        query = select(ExamRecord).where(
            ExamRecord.project_id == project_id,
            ExamRecord.student_id == student_id,
            ExamRecord.exam_name == exam_name,
            ExamRecord.subject == subject,
        )
        if exam_date is not None:
            query = query.where(ExamRecord.exam_date == exam_date)
        result = self.db.execute(query)
        return result.scalar_one_or_none()

    def _calculate_grade(self, marks_obtained: Decimal, max_marks: Decimal) -> str:
        """Calculate grade based on percentage."""
        if max_marks == 0:
            return "N/A"
        percentage = (marks_obtained / max_marks) * 100
        if percentage >= 90:
            return "A+"
        elif percentage >= 80:
            return "A"
        elif percentage >= 70:
            return "B+"
        elif percentage >= 60:
            return "B"
        elif percentage >= 50:
            return "C+"
        elif percentage >= 40:
            return "C"
        elif percentage >= 33:
            return "D"
        else:
            return "F"

    def get_record(
        self,
        record_id: int,
        project_id: int,
    ) -> ExamRecord:
        """Get exam record by ID."""
        result = self.db.execute(
            select(ExamRecord).where(
                ExamRecord.id == record_id,
                ExamRecord.project_id == project_id,
            )
        )
        record = result.scalar_one_or_none()
        if not record:
            raise NotFoundError("Exam record", str(record_id))
        return record

    def list_records(
        self,
        project_id: int,
        filters: ExamFilter | None = None,
        page: int = 1,
        page_size: int = 50,
    ) -> tuple[list[ExamRecordResponse], int]:
        """List exam records with filtering."""
        query = select(ExamRecord).where(ExamRecord.project_id == project_id)

        if filters:
            if filters.student_id:
                query = query.where(ExamRecord.student_id == filters.student_id)
            if filters.exam_name:
                query = query.where(ExamRecord.exam_name == filters.exam_name)
            if filters.subject:
                query = query.where(ExamRecord.subject == filters.subject)
            if filters.date_from:
                query = query.where(ExamRecord.exam_date >= filters.date_from)
            if filters.date_to:
                query = query.where(ExamRecord.exam_date <= filters.date_to)
            if filters.month and filters.year:
                # Filter by month and year
                start_date = date(filters.year, filters.month, 1)
                if filters.month == 12:
                    end_date = date(filters.year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = date(filters.year, filters.month + 1, 1) - timedelta(days=1)
                query = query.where(ExamRecord.exam_date >= start_date)
                query = query.where(ExamRecord.exam_date <= end_date)
            if filters.class_section:
                # Parse class_section like "3-A" into class_name and section
                class_name, section = self._parse_class_section(filters.class_section)
                query = query.join(Student).where(Student.class_name == class_name)
                if section:
                    query = query.where(Student.section == section)
            elif filters.class_name:
                query = query.join(Student).where(Student.class_name == filters.class_name)
                if filters.section:
                    query = query.where(Student.section == filters.section)

        # Count total
        count_result = self.db.execute(
            select(func.count()).select_from(query.subquery())
        )
        total = count_result.scalar() or 0

        # Apply pagination and ordering
        query = (
            query
            .order_by(ExamRecord.exam_date.desc(), ExamRecord.exam_name)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )

        result = self.db.execute(query)
        records = result.scalars().all()

        return [ExamRecordResponse.model_validate(self._record_to_response(r)) for r in records], total

    def update_record(
        self,
        record_id: int,
        project_id: int,
        request: ExamRecordUpdate,
    ) -> ExamRecordResponse:
        """Update an exam record."""
        record = self.get_record(record_id, project_id)

        # Get max_marks for validation
        max_marks = request.max_marks if request.max_marks is not None else record.max_marks

        # Validate marks if updating
        if request.marks_obtained is not None:
            if request.marks_obtained > max_marks:
                raise ValidationError(
                    f"marks_obtained ({request.marks_obtained}) exceeds max_marks ({max_marks})"
                )

        update_data = request.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(record, field, value)

        # Recalculate grade if marks changed
        if request.marks_obtained is not None and request.grade is None:
            record.grade = self._calculate_grade(record.marks_obtained, record.max_marks)

        self.db.flush()
        self.db.refresh(record)

        return ExamRecordResponse.model_validate(self._record_to_response(record))

    def delete_record(
        self,
        record_id: int,
        project_id: int,
    ) -> None:
        """Delete an exam record."""
        record = self.get_record(record_id, project_id)
        self.db.delete(record)
        self.db.flush()

    def get_exam_summary(
        self,
        project_id: int,
        exam_name: str,
        subject: str,
    ) -> ExamSummary:
        """Get summary statistics for an exam."""
        # Assuming 40% is pass mark
        pass_percentage = Decimal("0.4")

        query = select(
            func.count().label("total"),
            func.avg(ExamRecord.marks_obtained).label("avg"),
            func.max(ExamRecord.marks_obtained).label("highest"),
            func.min(ExamRecord.marks_obtained).label("lowest"),
        ).where(
            ExamRecord.project_id == project_id,
            ExamRecord.exam_name == exam_name,
            ExamRecord.subject == subject,
        )

        result = self.db.execute(query)
        row = result.one()

        # Count pass/fail
        pass_count_result = self.db.execute(
            select(func.count()).where(
                ExamRecord.project_id == project_id,
                ExamRecord.exam_name == exam_name,
                ExamRecord.subject == subject,
                ExamRecord.marks_obtained >= ExamRecord.max_marks * pass_percentage,
            )
        )
        pass_count = pass_count_result.scalar() or 0

        total = row.total or 0
        fail_count = total - pass_count

        return ExamSummary(
            exam_name=exam_name,
            subject=subject,
            total_students=total,
            average_marks=Decimal(str(row.avg or 0)).quantize(Decimal("0.01")),
            highest_marks=row.highest or Decimal("0"),
            lowest_marks=row.lowest or Decimal("0"),
            pass_count=pass_count,
            fail_count=fail_count,
        )

    # ==========================================
    # Bulk Operations
    # ==========================================

    def bulk_create_or_update(
        self,
        project_id: int,
        request: BulkExamCreate,
    ) -> BulkExamResponse:
        """Create or update exam records in bulk for a class."""
        errors = []
        successful = 0
        failed = 0

        # Parse class section
        class_name, section = self._parse_class_section(request.class_section)

        # Get all students for the class
        students = self._get_students_by_class(project_id, class_name, section)
        student_ids = {s.id for s in students}

        # Validate all student IDs before any DB operations
        for record in request.records:
            if record.student_id not in student_ids:
                errors.append({
                    "student_id": record.student_id,
                    "message": f"Student ID {record.student_id} not found in class {request.class_section}",
                })
                failed += 1

        if errors:
            return BulkExamResponse(
                total_records=len(request.records),
                successful=0,
                failed=failed,
                errors=errors,
                message="Validation failed. No records were saved.",
            )

        # All validations passed, proceed with DB operations
        for record in request.records:
            try:
                # Validate marks
                if record.marks_obtained > request.max_marks:
                    errors.append({
                        "student_id": record.student_id,
                        "message": f"Marks obtained ({record.marks_obtained}) exceeds max marks ({request.max_marks})",
                    })
                    failed += 1
                    continue

                existing = self._get_existing_record(
                    project_id, record.student_id, request.exam_name, request.subject, request.exam_date
                )

                # Calculate grade if not provided
                grade = record.grade
                if not grade:
                    grade = self._calculate_grade(record.marks_obtained, request.max_marks)

                if existing:
                    # Update existing record
                    existing.marks_obtained = record.marks_obtained
                    existing.max_marks = request.max_marks
                    existing.exam_date = request.exam_date
                    existing.grade = grade
                    existing.remarks = record.remarks
                else:
                    # Create new record
                    new_record = ExamRecord(
                        project_id=project_id,
                        student_id=record.student_id,
                        exam_name=request.exam_name,
                        subject=request.subject,
                        exam_date=request.exam_date,
                        max_marks=request.max_marks,
                        marks_obtained=record.marks_obtained,
                        grade=grade,
                        remarks=record.remarks,
                    )
                    self.db.add(new_record)

                successful += 1
            except Exception as e:
                errors.append({
                    "student_id": record.student_id,
                    "message": str(e),
                })
                failed += 1

        self.db.flush()

        return BulkExamResponse(
            total_records=len(request.records),
            successful=successful,
            failed=failed,
            errors=errors,
            message=f"Successfully saved {successful} exam records.",
        )

    def get_exam_by_class(
        self,
        project_id: int,
        class_section: str,
        exam_name: str,
        subject: str,
    ) -> ExamByClassResponse:
        """Get all student exam records for a class for a specific exam/subject."""
        class_name, section = self._parse_class_section(class_section)

        # Get all students in the class
        students = self._get_students_by_class(project_id, class_name, section)

        # Get existing exam records
        result = self.db.execute(
            select(ExamRecord).where(
                ExamRecord.project_id == project_id,
                ExamRecord.exam_name == exam_name,
                ExamRecord.subject == subject,
                ExamRecord.student_id.in_([s.id for s in students]),
            )
        )
        records = {r.student_id: r for r in result.scalars().all()}

        # Build response with all students
        student_data = []
        marks_list = []
        max_marks_value = None
        exam_date_value = None

        for student in students:
            record = records.get(student.id)
            if record:
                marks_list.append(record.marks_obtained)
                max_marks_value = record.max_marks
                exam_date_value = record.exam_date

            student_data.append(StudentExamEntry(
                student_id=student.id,
                student_name=student.student_name,
                class_name=student.class_name,
                section=student.section,
                marks_obtained=record.marks_obtained if record else None,
                max_marks=record.max_marks if record else None,
                grade=record.grade if record else None,
                remarks=record.remarks if record else None,
                record_id=record.id if record else None,
            ))

        # Calculate statistics
        average = sum(marks_list) / len(marks_list) if marks_list else None
        highest = max(marks_list) if marks_list else None
        lowest = min(marks_list) if marks_list else None

        return ExamByClassResponse(
            class_section=class_section,
            exam_name=exam_name,
            subject=subject,
            exam_date=exam_date_value,
            max_marks=max_marks_value,
            students=student_data,
            total_students=len(students),
            average_marks=Decimal(str(average)).quantize(Decimal("0.01")) if average else None,
            highest_marks=highest,
            lowest_marks=lowest,
        )

    def get_exam_names(self, project_id: int) -> list[str]:
        """Get distinct exam names for a project."""
        result = self.db.execute(
            select(ExamRecord.exam_name)
            .where(ExamRecord.project_id == project_id)
            .distinct()
            .order_by(ExamRecord.exam_name)
        )
        return [r[0] for r in result.all()]

    def get_subjects(self) -> list[str]:
        """Get list of available subjects."""
        return SUBJECTS

    def get_class_sections(self, project_id: int) -> list[dict]:
        """Get distinct class-section combinations for a project."""
        result = self.db.execute(
            select(Student.class_name, Student.section, func.count(Student.id).label("count"))
            .where(Student.project_id == project_id)
            .group_by(Student.class_name, Student.section)
            .order_by(Student.class_name, Student.section)
        )
        rows = result.all()
        return [
            {
                "class_name": r[0],
                "section": r[1],
                "class_section": f"{r[0]}-{r[1]}" if r[1] else r[0],
                "student_count": r[2],
            }
            for r in rows
        ]

    # ==========================================
    # Template Generation
    # ==========================================

    def generate_template(
        self,
        project_id: int,
        class_section: str | None = None,
        subject: str | None = None,
        month: int | None = None,
        year: int | None = None,
    ) -> bytes:
        """Generate Excel template for exam upload.

        Creates a properly formatted template with all required columns.
        When class_section is provided, pre-fills student data.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Exam Records"

        # Default to current month/year
        today = date.today()
        month = month or today.month
        year = year or today.year
        month_name = calendar.month_name[month]

        # Styles
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subheader_fill = PatternFill(start_color="8FAADC", end_color="8FAADC", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # Title row
        title_text = f"Exam Records - {month_name} {year}"
        if class_section:
            title_text = f"{class_section} - {title_text}"
        if subject:
            title_text += f" - {subject}"
        
        ws.merge_cells('A1:I1')
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

        # Headers - always include ALL required columns
        headers = [
            "Student Name",
            "Grade",
            "Exam Name", 
            "Subject",
            "Exam Date",
            "Max Marks",
            "Marks Obtained",
            "Grade (Auto)",
            "Remarks"
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Get students if class is specified
        start_row = 3
        if class_section:
            class_name, section = self._parse_class_section(class_section)
            students = self._get_students_by_class(project_id, class_name, section)

            default_exam_name = f"{month_name} {year} Exam"
            default_exam_date = f"{year}-{month:02d}-01"
            default_subject = subject if subject else ""

            for row_idx, student in enumerate(students, start=start_row):
                grade_str = f"{student.class_name}-{student.section}" if student.section else student.class_name
                
                ws.cell(row=row_idx, column=1, value=student.student_name).border = thin_border
                ws.cell(row=row_idx, column=2, value=grade_str).border = thin_border
                ws.cell(row=row_idx, column=3, value=default_exam_name).border = thin_border
                ws.cell(row=row_idx, column=4, value=default_subject).border = thin_border
                ws.cell(row=row_idx, column=5, value=default_exam_date).border = thin_border
                ws.cell(row=row_idx, column=6, value=100).border = thin_border  # Default max marks
                ws.cell(row=row_idx, column=7, value="").border = thin_border  # Marks obtained - to be filled
                ws.cell(row=row_idx, column=8, value="").border = thin_border  # Grade - auto calculated
                ws.cell(row=row_idx, column=9, value="").border = thin_border  # Remarks
        else:
            # Add sample rows
            sample_data = [
                ["John Doe", "10-A", f"{month_name} {year} Exam", "Mathematics", f"{year}-{month:02d}-15", 100, 85, "", "Good"],
                ["Jane Smith", "10-A", f"{month_name} {year} Exam", "Mathematics", f"{year}-{month:02d}-15", 100, 92, "", "Excellent"],
            ]
            for row_offset, row_data in enumerate(sample_data):
                row_idx = start_row + row_offset
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border

        # Column widths
        column_widths = {
            'A': 25,  # Student Name
            'B': 10,  # Grade
            'C': 25,  # Exam Name
            'D': 15,  # Subject
            'E': 12,  # Exam Date
            'F': 12,  # Max Marks
            'G': 15,  # Marks Obtained
            'H': 12,  # Grade (Auto)
            'I': 20,  # Remarks
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Add instructions sheet
        instructions_ws = wb.create_sheet("Instructions")
        instructions_ws.column_dimensions['A'].width = 25
        instructions_ws.column_dimensions['B'].width = 60

        instructions = [
            ("EXAM TEMPLATE INSTRUCTIONS", ""),
            ("", ""),
            ("REQUIRED COLUMNS:", ""),
            ("Student Name", "Must match exactly with student name in system"),
            ("Grade", "Class-Section format like '10-A' or '5-B'"),
            ("Exam Name", "Name of the exam (e.g., 'Mid-Term 2026', 'Unit Test 1')"),
            ("Subject", "Subject name - MUST be one of the valid subjects listed below"),
            ("Exam Date", "Date in YYYY-MM-DD format (e.g., 2026-01-15)"),
            ("Max Marks", "Maximum marks for the exam (e.g., 100)"),
            ("Marks Obtained", "Marks scored by the student (required)"),
            ("", ""),
            ("OPTIONAL COLUMNS:", ""),
            ("Grade (Auto)", "Leave empty - will be auto-calculated based on percentage"),
            ("Remarks", "Optional comments about student performance"),
            ("", ""),
            ("VALID SUBJECTS:", ""),
        ]
        
        for row_idx, (col1, col2) in enumerate(instructions, start=1):
            cell1 = instructions_ws.cell(row=row_idx, column=1, value=col1)
            cell2 = instructions_ws.cell(row=row_idx, column=2, value=col2)
            if row_idx == 1:
                cell1.font = Font(bold=True, size=14)
            elif col1 and col1.endswith(":"):
                cell1.font = Font(bold=True)

        # List all valid subjects
        subject_start_row = len(instructions) + 1
        for idx, subj in enumerate(SUBJECTS):
            instructions_ws.cell(row=subject_start_row + idx, column=1, value=f"{idx + 1}.")
            instructions_ws.cell(row=subject_start_row + idx, column=2, value=subj)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # ==========================================
    # Excel Upload Processing
    # ==========================================

    def process_excel_upload(
        self,
        project_id: int,
        file_content: bytes,
        upload_id: int | None = None,
    ) -> ExamUploadResult:
        """Process exam Excel upload with validation."""
        logger.info(f"[EXAM UPLOAD] Starting - project_id={project_id}, file_size={len(file_content)} bytes")

        try:
            wb = load_workbook(BytesIO(file_content), data_only=True)
            ws = wb.active
        except Exception as e:
            logger.error(f"[EXAM UPLOAD] Failed to load Excel: {str(e)}")
            raise ValidationError(f"Invalid Excel file: {str(e)}")

        errors: list[ExamUploadError] = []
        successful_rows = 0
        failed_rows = 0
        skipped_rows = 0

        # Parse headers
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        logger.info(f"[EXAM UPLOAD] Headers: {headers}")

        # Get all students in project indexed by name (case-insensitive)
        students_result = self.db.execute(
            select(Student).where(Student.project_id == project_id)
        )
        students_by_name: dict[str, Student] = {}
        for s in students_result.scalars().all():
            key = s.student_name.lower().strip()
            students_by_name[key] = s
        logger.info(f"[EXAM UPLOAD] Found {len(students_by_name)} students in database")

        # Map header names to column indices
        col_map = {
            "student_name": None,
            "grade": None,
            "exam_name": None,
            "subject": None,
            "exam_date": None,
            "max_marks": None,
            "marks_obtained": None,
            "grade_col": None,
            "remarks": None,
        }

        for idx, header in enumerate(headers):
            if "student" in header and "name" in header:
                col_map["student_name"] = idx
            elif "grade" in header and "class" in header:
                col_map["grade"] = idx
            elif header == "grade":
                if col_map["grade"] is None:
                    col_map["grade"] = idx
                else:
                    col_map["grade_col"] = idx
            elif "exam" in header and "name" in header:
                col_map["exam_name"] = idx
            elif "subject" in header:
                col_map["subject"] = idx
            elif "date" in header:
                col_map["exam_date"] = idx
            elif "max" in header:
                col_map["max_marks"] = idx
            elif "obtained" in header or "marks" in header:
                if col_map["max_marks"] is not None and col_map["marks_obtained"] is None:
                    col_map["marks_obtained"] = idx
                elif col_map["marks_obtained"] is None:
                    col_map["marks_obtained"] = idx
            elif "remark" in header:
                col_map["remarks"] = idx

        logger.info(f"[EXAM UPLOAD] Column mapping: {col_map}")

        # Process rows
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                skipped_rows += 1
                continue

            try:
                # Extract values
                student_name = str(row[col_map["student_name"]]).strip() if col_map["student_name"] is not None and row[col_map["student_name"]] else None
                grade_class = str(row[col_map["grade"]]).strip() if col_map["grade"] is not None and row[col_map["grade"]] else None
                exam_name = str(row[col_map["exam_name"]]).strip() if col_map["exam_name"] is not None and row[col_map["exam_name"]] else None
                subject = str(row[col_map["subject"]]).strip() if col_map["subject"] is not None and row[col_map["subject"]] else None
                exam_date_str = str(row[col_map["exam_date"]]).strip() if col_map["exam_date"] is not None and row[col_map["exam_date"]] else None
                max_marks_str = str(row[col_map["max_marks"]]).strip() if col_map["max_marks"] is not None and row[col_map["max_marks"]] else None
                marks_str = str(row[col_map["marks_obtained"]]).strip() if col_map["marks_obtained"] is not None and row[col_map["marks_obtained"]] else None
                grade_value = str(row[col_map["grade_col"]]).strip() if col_map["grade_col"] is not None and row[col_map["grade_col"]] else None
                remarks = str(row[col_map["remarks"]]).strip() if col_map["remarks"] is not None and row[col_map["remarks"]] else None

                if not student_name:
                    skipped_rows += 1
                    continue

                # Validate student
                student_key = student_name.lower()
                student = students_by_name.get(student_key)
                if not student:
                    errors.append(ExamUploadError(
                        row=row_num,
                        student_name=student_name,
                        message=f"Student '{student_name}' not found in system",
                    ))
                    failed_rows += 1
                    continue

                # Validate required fields
                if not exam_name:
                    errors.append(ExamUploadError(
                        row=row_num,
                        student_name=student_name,
                        column="Exam Name",
                        message="Exam name is required",
                    ))
                    failed_rows += 1
                    continue

                if not subject:
                    errors.append(ExamUploadError(
                        row=row_num,
                        student_name=student_name,
                        column="Subject",
                        message="Subject is required",
                    ))
                    failed_rows += 1
                    continue

                # Parse exam date
                exam_date = None
                if exam_date_str:
                    try:
                        if "-" in exam_date_str:
                            parts = exam_date_str.split("-")
                            exam_date = date(int(parts[0]), int(parts[1]), int(parts[2].split()[0]))
                        else:
                            exam_date = date.today()
                    except Exception:
                        exam_date = date.today()
                else:
                    exam_date = date.today()

                # Parse max marks
                try:
                    max_marks = Decimal(max_marks_str) if max_marks_str else Decimal("100")
                except Exception:
                    max_marks = Decimal("100")

                # Parse marks obtained
                if not marks_str:
                    errors.append(ExamUploadError(
                        row=row_num,
                        student_name=student_name,
                        column="Marks Obtained",
                        message="Marks obtained is required",
                    ))
                    failed_rows += 1
                    continue

                try:
                    marks_obtained = Decimal(marks_str)
                except Exception:
                    errors.append(ExamUploadError(
                        row=row_num,
                        student_name=student_name,
                        column="Marks Obtained",
                        message=f"Invalid marks value: '{marks_str}'",
                    ))
                    failed_rows += 1
                    continue

                # Validate marks
                if marks_obtained > max_marks:
                    errors.append(ExamUploadError(
                        row=row_num,
                        student_name=student_name,
                        message=f"Marks obtained ({marks_obtained}) exceeds max marks ({max_marks})",
                    ))
                    failed_rows += 1
                    continue

                # Calculate grade if not provided
                grade = grade_value if grade_value and grade_value.lower() != "none" else None
                if not grade:
                    grade = self._calculate_grade(marks_obtained, max_marks)

                # Check for existing record and update or create
                # Uniqueness includes exam_date, so same exam type on different dates creates new records
                existing = self._get_existing_record(
                    project_id, student.id, exam_name, subject, exam_date
                )

                if existing:
                    existing.marks_obtained = marks_obtained
                    existing.max_marks = max_marks
                    existing.exam_date = exam_date
                    existing.grade = grade
                    existing.remarks = remarks if remarks and remarks.lower() != "none" else None
                    if upload_id:
                        existing.upload_id = upload_id
                else:
                    record = ExamRecord(
                        project_id=project_id,
                        student_id=student.id,
                        exam_name=exam_name,
                        subject=subject,
                        exam_date=exam_date,
                        max_marks=max_marks,
                        marks_obtained=marks_obtained,
                        grade=grade,
                        remarks=remarks if remarks and remarks.lower() != "none" else None,
                        upload_id=upload_id,
                    )
                    self.db.add(record)

                successful_rows += 1

            except Exception as e:
                errors.append(ExamUploadError(
                    row=row_num,
                    message=f"Error processing row: {str(e)}",
                ))
                failed_rows += 1

        self.db.flush()

        total = successful_rows + failed_rows + skipped_rows
        logger.info(f"[EXAM UPLOAD] Completed: {successful_rows} OK, {failed_rows} failed, {skipped_rows} skipped")

        return ExamUploadResult(
            total_rows=total,
            successful_rows=successful_rows,
            failed_rows=failed_rows,
            skipped_rows=skipped_rows,
            errors=errors,
            message=f"Processed {successful_rows} exam records successfully.",
        )

    # ==========================================
    # Helper Methods
    # ==========================================

    def _parse_class_section(self, class_section: str) -> tuple[str, str | None]:
        """Parse class-section string like '3-A' into (class_name, section)."""
        if "-" in class_section:
            parts = class_section.rsplit("-", 1)
            return parts[0], parts[1]
        return class_section, None

    def _get_students_by_class(
        self, project_id: int, class_name: str, section: str | None = None
    ) -> list[Student]:
        """Get all students in a class, optionally filtered by section."""
        query = select(Student).where(
            Student.project_id == project_id,
            Student.class_name == class_name,
        )
        if section:
            query = query.where(Student.section == section)

        query = query.order_by(Student.student_name)
        result = self.db.execute(query)
        return list(result.scalars().all())
