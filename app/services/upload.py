"""Upload processing service for Excel files."""

import logging
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.exceptions import UploadError, ValidationError
from app.models.attendance import AttendanceRecord, AttendanceStatus
from app.models.exam import ExamRecord
from app.models.upload import Upload, UploadError as UploadErrorModel, UploadStatus, UploadType
from app.schemas.upload import UploadErrorResponse, UploadResult

# Setup debug logger
logger = logging.getLogger(__name__)


class UploadService:
    """Excel upload processing service."""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def process_attendance_upload(
        self,
        project_id: UUID,
        user_id: UUID,
        file_content: bytes,
        file_name: str,
    ) -> UploadResult:
        """
        Process attendance Excel upload.
        Allows partial success - invalid rows are skipped.
        """
        logger.info(f"[ATTENDANCE UPLOAD] Starting upload for file: {file_name}, project: {project_id}, user: {user_id}")
        logger.debug(f"[ATTENDANCE UPLOAD] File size: {len(file_content)} bytes")
        
        # Create upload record
        upload = Upload(
            project_id=project_id,
            upload_type=UploadType.ATTENDANCE,
            file_name=file_name,
            file_size=len(file_content),
            status=UploadStatus.PROCESSING,
            uploaded_by_id=user_id,
            processing_started_at=datetime.now(timezone.utc),
        )
        self.db.add(upload)
        await self.db.flush()
        logger.debug(f"[ATTENDANCE UPLOAD] Created upload record with ID: {upload.id}")

        try:
            # Parse Excel file
            rows = self._parse_excel(file_content)
            upload.total_rows = len(rows)
            logger.info(f"[ATTENDANCE UPLOAD] Parsed {len(rows)} data rows from Excel")

            errors: list[UploadErrorModel] = []
            successful = 0

            for row_num, row in enumerate(rows, start=2):  # Start at 2 (header is 1)
                logger.debug(f"[ATTENDANCE UPLOAD] Processing row {row_num}: {row}")
                try:
                    record = self._validate_attendance_row(row, row_num, project_id, upload.id)
                    self.db.add(record)
                    successful += 1
                    logger.debug(f"[ATTENDANCE UPLOAD] Row {row_num} SUCCESS - Created record: student_id={record.student_id}, student_name={record.student_name}, date={record.attendance_date}, status={record.status}")
                except ValidationError as e:
                    logger.warning(f"[ATTENDANCE UPLOAD] Row {row_num} FAILED - {e.message} (column={e.details.get('column')}, value={e.details.get('value')})")
                    error = UploadErrorModel(
                        upload_id=upload.id,
                        row_number=row_num,
                        column_name=e.details.get("column"),
                        error_type="VALIDATION_ERROR",
                        error_message=e.message,
                        raw_value=str(e.details.get("value", "")),
                    )
                    errors.append(error)
                    self.db.add(error)

            upload.successful_rows = successful
            upload.failed_rows = len(errors)
            upload.status = (
                UploadStatus.SUCCESS if not errors
                else UploadStatus.PARTIAL if successful > 0
                else UploadStatus.FAILED
            )
            upload.processing_completed_at = datetime.now(timezone.utc)

            logger.info(f"[ATTENDANCE UPLOAD] Upload complete - Status: {upload.status.value}, Successful: {successful}, Failed: {len(errors)}, Total: {upload.total_rows}")
            
            await self.db.flush()
            logger.debug(f"[ATTENDANCE UPLOAD] Database flush completed - records should be persisted")

            return UploadResult(
                upload_id=upload.id,
                status=upload.status,
                total_rows=upload.total_rows,
                successful_rows=upload.successful_rows,
                failed_rows=upload.failed_rows,
                errors=[
                    UploadErrorResponse(
                        id=e.id,
                        upload_id=e.upload_id,
                        row_number=e.row_number,
                        column_name=e.column_name,
                        error_type=e.error_type,
                        error_message=e.error_message,
                        raw_value=e.raw_value,
                    )
                    for e in errors
                ],
                message=self._get_result_message(upload),
            )

        except Exception as e:
            upload.status = UploadStatus.FAILED
            upload.error_message = str(e)
            upload.processing_completed_at = datetime.now(timezone.utc)
            await self.db.flush()
            raise UploadError(f"Failed to process attendance upload: {str(e)}")

    async def process_exam_upload(
        self,
        project_id: UUID,
        user_id: UUID,
        file_content: bytes,
        file_name: str,
    ) -> UploadResult:
        """
        Process exam Excel upload.
        STRICT: Any invalid row triggers FULL ROLLBACK.
        """
        # Create upload record
        upload = Upload(
            project_id=project_id,
            upload_type=UploadType.EXAM,
            file_name=file_name,
            file_size=len(file_content),
            status=UploadStatus.PROCESSING,
            uploaded_by_id=user_id,
            processing_started_at=datetime.now(timezone.utc),
        )
        self.db.add(upload)
        await self.db.flush()

        try:
            # Parse Excel file
            rows = self._parse_excel(file_content)
            upload.total_rows = len(rows)

            # First pass: validate ALL rows before inserting any
            validated_records: list[ExamRecord] = []
            errors: list[UploadErrorModel] = []

            for row_num, row in enumerate(rows, start=2):
                try:
                    record = self._validate_exam_row(row, row_num, project_id, upload.id)
                    validated_records.append(record)
                except ValidationError as e:
                    error = UploadErrorModel(
                        upload_id=upload.id,
                        row_number=row_num,
                        column_name=e.details.get("column"),
                        error_type="VALIDATION_ERROR",
                        error_message=e.message,
                        raw_value=str(e.details.get("value", "")),
                    )
                    errors.append(error)

            # If ANY errors, rollback and fail
            if errors:
                for error in errors:
                    self.db.add(error)

                upload.successful_rows = 0
                upload.failed_rows = len(errors)
                upload.status = UploadStatus.FAILED
                upload.error_message = f"Validation failed for {len(errors)} rows. Full rollback applied."
                upload.processing_completed_at = datetime.now(timezone.utc)

                await self.db.flush()

                return UploadResult(
                    upload_id=upload.id,
                    status=upload.status,
                    total_rows=upload.total_rows,
                    successful_rows=0,
                    failed_rows=len(errors),
                    errors=[
                        UploadErrorResponse(
                            id=e.id,
                            upload_id=e.upload_id,
                            row_number=e.row_number,
                            column_name=e.column_name,
                            error_type=e.error_type,
                            error_message=e.error_message,
                            raw_value=e.raw_value,
                        )
                        for e in errors
                    ],
                    message="Exam upload failed. All rows rejected due to validation errors.",
                )

            # All valid, insert all records
            for record in validated_records:
                self.db.add(record)

            upload.successful_rows = len(validated_records)
            upload.failed_rows = 0
            upload.status = UploadStatus.SUCCESS
            upload.processing_completed_at = datetime.now(timezone.utc)

            await self.db.flush()

            return UploadResult(
                upload_id=upload.id,
                status=upload.status,
                total_rows=upload.total_rows,
                successful_rows=upload.successful_rows,
                failed_rows=0,
                errors=[],
                message=f"Successfully imported {upload.successful_rows} exam records.",
            )

        except Exception as e:
            upload.status = UploadStatus.FAILED
            upload.error_message = str(e)
            upload.processing_completed_at = datetime.now(timezone.utc)
            await self.db.flush()
            raise UploadError(f"Failed to process exam upload: {str(e)}")

    def _parse_excel(self, file_content: bytes) -> list[dict[str, Any]]:
        """Parse Excel file and return list of row dictionaries."""
        try:
            workbook = load_workbook(filename=BytesIO(file_content), read_only=True)
            sheet = workbook.active

            if sheet is None:
                raise UploadError("Excel file has no active sheet")

            rows = list(sheet.iter_rows(values_only=True))
            logger.debug(f"[EXCEL PARSE] Total raw rows in Excel (including header): {len(rows)}")
            
            if len(rows) < 2:
                raise UploadError("Excel file must have a header row and at least one data row")

            # First row is headers
            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            logger.debug(f"[EXCEL PARSE] Detected headers: {headers}")

            # Convert remaining rows to dictionaries
            data = []
            skipped_empty_rows = 0
            for row_idx, row in enumerate(rows[1:], start=2):
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_dict[headers[i]] = value
                if any(row_dict.values()):  # Skip empty rows
                    data.append(row_dict)
                    logger.debug(f"[EXCEL PARSE] Row {row_idx} extracted: {row_dict}")
                else:
                    skipped_empty_rows += 1
                    logger.debug(f"[EXCEL PARSE] Row {row_idx} SKIPPED (all values empty): {row}")

            logger.info(f"[EXCEL PARSE] Summary: {len(data)} data rows extracted, {skipped_empty_rows} empty rows skipped")
            return data

        except Exception as e:
            if isinstance(e, UploadError):
                raise
            raise UploadError(f"Failed to parse Excel file: {str(e)}")

    def _validate_attendance_row(
        self,
        row: dict[str, Any],
        row_num: int,
        project_id: UUID,
        upload_id: UUID,
    ) -> AttendanceRecord:
        """Validate and create attendance record from row."""
        logger.debug(f"[VALIDATE ROW {row_num}] Input data: {row}")
        
        # Required fields
        student_id = row.get("student_id")
        logger.debug(f"[VALIDATE ROW {row_num}] student_id raw value: {student_id!r} (type: {type(student_id).__name__})")
        if not student_id:
            raise ValidationError(
                "Student ID is required",
                details={"column": "student_id", "row": row_num},
            )

        student_name = row.get("student_name")
        logger.debug(f"[VALIDATE ROW {row_num}] student_name raw value: {student_name!r} (type: {type(student_name).__name__})")
        if not student_name:
            raise ValidationError(
                "Student name is required",
                details={"column": "student_name", "row": row_num},
            )

        # Parse date
        date_value = row.get("attendance_date")
        logger.debug(f"[VALIDATE ROW {row_num}] attendance_date raw value: {date_value!r} (type: {type(date_value).__name__})")
        if not date_value:
            raise ValidationError(
                "Attendance date is required",
                details={"column": "attendance_date", "row": row_num},
            )

        try:
            if isinstance(date_value, datetime):
                attendance_date = date_value.date()
                logger.debug(f"[VALIDATE ROW {row_num}] Parsed datetime -> date: {attendance_date}")
            elif isinstance(date_value, str):
                from datetime import date
                # Try common formats
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]:
                    try:
                        attendance_date = datetime.strptime(date_value, fmt).date()
                        logger.debug(f"[VALIDATE ROW {row_num}] Parsed string with format '{fmt}' -> date: {attendance_date}")
                        break
                    except ValueError:
                        continue
                else:
                    raise ValueError("Unknown date format")
            else:
                attendance_date = date_value
                logger.debug(f"[VALIDATE ROW {row_num}] Using date value as-is: {attendance_date} (type: {type(date_value).__name__})")
        except Exception:
            raise ValidationError(
                f"Invalid date format: {date_value}",
                details={"column": "attendance_date", "row": row_num, "value": str(date_value)},
            )

        # Parse status
        status_value = str(row.get("status", "")).lower().strip()
        logger.debug(f"[VALIDATE ROW {row_num}] status raw value: {row.get('status')!r} -> normalized: '{status_value}'")
        try:
            status = AttendanceStatus(status_value)
        except ValueError:
            raise ValidationError(
                f"Invalid status: {status_value}. Must be one of: present, absent, late, excused",
                details={"column": "status", "row": row_num, "value": status_value},
            )

        remarks = str(row.get("remarks", "")).strip() if row.get("remarks") else None
        logger.debug(f"[VALIDATE ROW {row_num}] remarks: {remarks!r}")
        
        record = AttendanceRecord(
            project_id=project_id,
            student_id=str(student_id).strip(),
            student_name=str(student_name).strip(),
            attendance_date=attendance_date,
            status=status,
            remarks=remarks,
            upload_id=upload_id,
        )
        
        logger.debug(f"[VALIDATE ROW {row_num}] Created AttendanceRecord: student_id={record.student_id}, student_name={record.student_name}, date={record.attendance_date}, status={record.status}")
        return record

    def _validate_exam_row(
        self,
        row: dict[str, Any],
        row_num: int,
        project_id: UUID,
        upload_id: UUID,
    ) -> ExamRecord:
        """Validate and create exam record from row. Strict validation."""
        # Required fields
        student_id = row.get("student_id")
        if not student_id:
            raise ValidationError(
                "Student ID is required",
                details={"column": "student_id", "row": row_num},
            )

        student_name = row.get("student_name")
        if not student_name:
            raise ValidationError(
                "Student name is required",
                details={"column": "student_name", "row": row_num},
            )

        exam_name = row.get("exam_name")
        if not exam_name:
            raise ValidationError(
                "Exam name is required",
                details={"column": "exam_name", "row": row_num},
            )

        subject = row.get("subject")
        if not subject:
            raise ValidationError(
                "Subject is required",
                details={"column": "subject", "row": row_num},
            )

        # Parse date
        date_value = row.get("exam_date")
        if not date_value:
            raise ValidationError(
                "Exam date is required",
                details={"column": "exam_date", "row": row_num},
            )

        try:
            if isinstance(date_value, datetime):
                exam_date = date_value.date()
            elif isinstance(date_value, str):
                from datetime import date
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]:
                    try:
                        exam_date = datetime.strptime(date_value, fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    raise ValueError("Unknown date format")
            else:
                exam_date = date_value
        except Exception:
            raise ValidationError(
                f"Invalid date format: {date_value}",
                details={"column": "exam_date", "row": row_num, "value": str(date_value)},
            )

        # Parse marks
        try:
            max_marks = Decimal(str(row.get("max_marks", 0)))
            if max_marks <= 0:
                raise ValidationError(
                    "Max marks must be greater than 0",
                    details={"column": "max_marks", "row": row_num, "value": str(row.get("max_marks"))},
                )
        except (InvalidOperation, ValueError):
            raise ValidationError(
                f"Invalid max marks: {row.get('max_marks')}",
                details={"column": "max_marks", "row": row_num, "value": str(row.get("max_marks"))},
            )

        try:
            marks_obtained = Decimal(str(row.get("marks_obtained", 0)))
            if marks_obtained < 0:
                raise ValidationError(
                    "Marks obtained cannot be negative",
                    details={"column": "marks_obtained", "row": row_num, "value": str(row.get("marks_obtained"))},
                )
        except (InvalidOperation, ValueError):
            raise ValidationError(
                f"Invalid marks obtained: {row.get('marks_obtained')}",
                details={"column": "marks_obtained", "row": row_num, "value": str(row.get("marks_obtained"))},
            )

        # CRITICAL: marks_obtained must not exceed max_marks
        if marks_obtained > max_marks:
            raise ValidationError(
                f"marks_obtained ({marks_obtained}) exceeds max_marks ({max_marks})",
                details={
                    "column": "marks_obtained",
                    "row": row_num,
                    "value": str(marks_obtained),
                    "max_marks": str(max_marks),
                },
            )

        return ExamRecord(
            project_id=project_id,
            student_id=str(student_id).strip(),
            student_name=str(student_name).strip(),
            exam_name=str(exam_name).strip(),
            subject=str(subject).strip(),
            exam_date=exam_date,
            max_marks=max_marks,
            marks_obtained=marks_obtained,
            grade=str(row.get("grade", "")).strip() if row.get("grade") else None,
            remarks=str(row.get("remarks", "")).strip() if row.get("remarks") else None,
            upload_id=upload_id,
        )

    def _get_result_message(self, upload: Upload) -> str:
        """Generate result message for upload."""
        if upload.status == UploadStatus.SUCCESS:
            return f"Successfully imported {upload.successful_rows} records."
        elif upload.status == UploadStatus.PARTIAL:
            return (
                f"Partially imported: {upload.successful_rows} successful, "
                f"{upload.failed_rows} failed out of {upload.total_rows} rows."
            )
        else:
            return f"Upload failed: {upload.error_message or 'Unknown error'}"
